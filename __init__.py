import Member
import Product
import Orderhistory
import shelve
import Question
import Admin
import Voucher
import Supplier
import FeedbackSimpleDB

from News import News
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
import os
import datetime
from werkzeug.utils import secure_filename
from FeedbackSimpleDB import add_question, add_news
from Question import Question

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from ReportGeneration import *
from pyecharts import options as opts
from pyecharts.charts import Bar, Calendar, Pie, Liquid, Page, Tab
from markupsafe import Markup

from forms import CreateMemberForm, CreateProductForm, CreateQuestionForm, CreateLoginForm, CreateCardForm, \
    CreateAdminForm, CreateVoucherForm, VoucherForm, CreateSearchForm, CreateSupplierForm, CreateReplyForm, CreateNewsForm

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads/'
app.secret_key = 'PKfEKJh0'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg', 'gif'])


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def homepage():
    # db = shelve.open('database.db', 'c')
    # db['Question'] = {}
    # db.close()
    return render_template('homepage.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    create_login_form = CreateLoginForm(request.form)

    db = shelve.open('database.db', 'r')
    login_list = db["Members"]
    supplier_list = db['Supplier']
    admin_list = db["Admin"]

    if request.method == "POST" and create_login_form.validate():
        email = create_login_form.email.data
        password = create_login_form.password.data

        # Check for member login
        for member_id, member in login_list.items():
            if member.get_email() == email and member.get_password() == password:
                session['name'] = member.get_first_name() + " " + member.get_last_name()
                session['member_id'] = member_id
                flash("Login successful", "success")
                return redirect(url_for('outbox'))

        # Check for supplier login
        for supplier_id, supplier in supplier_list.items():
            if supplier.get_company_email() == email and supplier.get_password() == password:
                session['name'] = supplier.get_company_name()
                session['member_id'] = supplier_id
                session['supplier'] = "active"
                flash("Supplier login successful", "success")
                return redirect(url_for('outbox'))

        # Check for admin login
        for admin_id, admin in admin_list.items():
            if admin.get_email() == email and admin.get_password() == password:
                session['name'] = admin.get_first_name() + " " + admin.get_last_name()
                session['member_id'] = admin_id
                session['admin'] = "active"
                flash("Admin login successful", "success")
                return redirect(url_for('admin'))

        # If none of the login attempts are successful, display an error message
        flash("Invalid email or password. Please try again", "error")

    return render_template('login.html', form=create_login_form)


@app.route("/profile")
def profile():
    if "name" in session:
        supplier_info = None
        name = session["name"]
        id = session["member_id"]
        if 'admin' in session:
            admin = session['admin']
        else:
            admin = None

        if 'supplier' in session:
            db = shelve.open('database.db', 'r')
            supplier = session['supplier']
            supplier_dict = db['Supplier']
            id = session["member_id"]
            supplier_info = supplier_dict[id]
        else:
            supplier = None

        member_dict = {}
        db = shelve.open('database.db', 'r')
        member_dict = db['Members']
        user_info = member_dict[id]

        order_hist = db['OrderHist']
        member_orderhist = []
        for order_id in order_hist:
            if order_hist[order_id].get_name() == name:
                member_orderhist.append(order_hist[order_id])

        voucher_dict = db['Vouchers']
        voucher_list = []
        voucher_id = member_dict[id].get_vouchers()
        for i in voucher_id:
            voucher_list.append(voucher_dict[i])


    return render_template('profile.html', admin=admin, supplier=supplier,
                           name=name, id=id, user=user_info, supplier_user=supplier_info, history=member_orderhist,
                           vouchers=voucher_list)


@app.route("/logout")
def logout():
    session.pop('name', None)
    session.pop('member_id', None)
    if 'admin' in session:
        session.pop('admin', None)
    session['cart'] = []
    db = shelve.open('database.db', 'c')
    db['Outbox'] = {}
    db.close()
    return redirect(url_for('homepage'))


@app.route('/outbox')
def outbox():
    if request.method == 'POST':
        return redirect(url_for('filter_outbox'))
    # Retrieve products from inventory
    inventory_dict = {}
    db_inventory = shelve.open('database.db', 'r')
    inventory_dict = db_inventory['Inventory']
    db_inventory.close()

    categories = set(product.get_category() for product in inventory_dict.values())
    return render_template('outbox.html', outbox_products=inventory_dict.values(), categories=categories)


@app.route('/view_cart')
def view_cart():
    outbox_dict = {}
    db_outbox = shelve.open('database.db', 'r')
    outbox_dict = db_outbox['Outbox']
    db_outbox.close()

    outbox_list = list(outbox_dict.values())
    cart_list = session.get('cart', [])
    return render_template('viewcart.html', count=len(cart_list), outbox_list=outbox_list)


@app.route('/add_to_outbox/<int:product_id>')
def add_to_outbox(product_id):
    inventory_dict = {}
    db_inventory = shelve.open('database.db', 'r')
    inventory_dict = db_inventory['Inventory']
    db_inventory.close()

    selected_product = inventory_dict.get(product_id)

    outbox_dict = {}
    db_outbox = shelve.open('database.db', 'c')
    try:
        outbox_dict = db_outbox['Outbox']
    except:
        print("Error in retrieving products from database.db")

    if 'cart' not in session:
        session['cart'] = []

    if product_id in session['cart']:
        flash('Item already added to cart.')
    else:
        session['cart'].append(product_id)

    session.modified = True

    outbox_dict[selected_product.get_product_id()] = selected_product
    db_outbox['Outbox'] = outbox_dict
    db_outbox.close()
    flash("Item successfully added to cart.", "success")
    return redirect(url_for('outbox'))


@app.route('/deleteitem/<int:id>', methods=['POST'])
def delete_cart(id):
    cart_dict = {}
    db = shelve.open('database.db', 'w')
    cart_dict = db['Outbox']
    cart_dict.pop(id)
    db['Outbox'] = cart_dict
    db.close()
    cart_list = session.get('cart', [])
    if id in cart_list:
        cart_list.remove(id)

    session['cart'] = cart_list
    session.modified = True

    return redirect(url_for('view_cart'))


@app.route('/checkout', methods=['GET', 'POST'])
def checkout():
    cart_list = session.get('cart', [])
    if 'name' in session:
        name = session['name']
    else:
        name = None

    if 'admin' in session:
        flash('YOU ARE IN ADMIN MODE. CANNOT CHECK OUT')
        return redirect(url_for('view_cart'))

    if not cart_list:
        flash("Your cart is empty. Add items to your cart before proceeding to checkout.", "warning")
        return redirect(url_for('view_cart'))

    create_card_form = CreateCardForm(request.form)
    db = shelve.open("database.db", "c")
    checkout_dict = {}
    checkout_dict = db['Outbox']
    discount = None

    total_price = sum(float(item.get_price()) for item in checkout_dict.values())
    vouchers = []
    voucher_dict = []

    if 'name' in session and 'admin' not in session:
        voucher_dict = db['Vouchers']
        member_dict = db['Members']
        member_id = session['member_id']
        voucher_list = member_dict[member_id].get_vouchers()
        for i in voucher_list:
            vouchers.append(voucher_dict[i])
        print(vouchers)

    if request.method == "POST" and create_card_form.validate():
        order_dict = db['OrderHist']
        applied_voucher = None

        if 'name' in session:
            id = session['member_id']
            memberdb = db['Members']
            product = [item.get_name() for item in checkout_dict.values()]
            # date = datetime.date.today()

            # FOR FAKE ORDERHIST DATES
            from datetime import datetime
            date_str = '2024-02-11'
            date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Check if a voucher has been applied
            if 'applied_voucher' in session:
                applied_voucher_id = session['applied_voucher']
                applied_voucher_discount = voucher_dict[applied_voucher_id].get_discount()
                discount = float(applied_voucher_discount)
                # Apply voucher discount to the total amount
                total_price -= total_price * (discount / 100)

                member_dict = db['Members']
                voucher_list = member_dict[id].get_vouchers()

                if applied_voucher_id in voucher_list:
                    voucher_list.remove(applied_voucher_id)
                    member_dict[id].set_voucher_list(voucher_list)
                    db['Members'] = member_dict

            discount = int(discount) if discount is not None else None

            order_hist = Orderhistory.OrderHistory(
                session['name'], memberdb[id].get_email(),
                product, date, memberdb[id].get_phone(),
                total_price, discount
            )

            existing_ids = set(order_dict.keys())
            missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
            if missing_ids:
                my_key = min(missing_ids)
            else:
                my_key = len(order_dict) + 1

            order_hist.set_order_id(my_key)
            order_dict[my_key] = order_hist
            db['OrderHist'] = order_dict
        else:
            product = [item.get_name() for item in checkout_dict.values()]
            # date = datetime.date.today()

            # FOR FAKE ORDERHIST DATES
            from datetime import datetime
            date_str = '2024-02-11'
            date = datetime.strptime(date_str, '%Y-%m-%d').date()

            order_hist = Orderhistory.OrderHistory(
                "Guest", "Null", product, date, "Null",
                total_price, "None"
            )
            existing_ids = set(order_dict.keys())
            missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
            if missing_ids:
                my_key = min(missing_ids)
            else:
                my_key = len(order_dict) + 1

            existing_ids = set(order_dict.keys())
            missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
            if missing_ids:
                my_key = min(missing_ids)
            else:
                my_key = len(order_dict) + 1

            order_hist.set_order_id(my_key)
            order_dict[my_key] = order_hist
            db['OrderHist'] = order_dict

        for id in checkout_dict:
            if id in cart_list:
                cart_list.remove(id)

        session['cart'] = []
        if 'applied_voucher' in session:
            session.pop('applied_voucher', None)
        session.modified = True

        nothing = {}
        db['Outbox'] = nothing
        db.close()
        return render_template("checkoutconfirmation.html")
    db.close()
    return render_template('checkout.html',
                           cart_items=checkout_dict.values(), total_price=total_price, form=create_card_form,
                           vouchers=vouchers, name=name)

@app.route('/set_voucher_session/<voucher_id>', methods=['POST'])
def set_voucher_session(voucher_id):
    session['applied_voucher'] = voucher_id
    session.modified = True
    return jsonify({'success': True})


@app.route('/orderhistory')
def orderhistory():
    db = shelve.open("database.db", "c")
    order_dict = db['OrderHist']
    order_list = []

    for key in order_dict:
        orderid = order_dict.get(key)
        order_list.append(orderid)

    db.close()
    return render_template("orderhistory.html", count=len(order_dict), order_list=order_list)


@app.route('/deleteorder/<int:id>', methods=['POST'])
def delete_order(id):
    order_dict = {}
    db = shelve.open('database.db', 'w')
    order_dict = db['OrderHist']
    order_dict.pop(id)
    db['OrderHist'] = order_dict
    db.close()
    return redirect(url_for('orderhistory'))


@app.route('/admin')
def admin():
    if 'admin' in session:
        name = session['name']
        db = shelve.open('database.db', 'r')
        membercount = len(db['Members'])
        admincount = len(db['Admin'])
        suppliercount = len(db['Supplier'])
        inventorycount = len(db['Inventory'])
        ordercount = len(db['OrderHist'])
        foremcount = len(db['Question'])
        newscount = len(db['News'])
        vouchercount = len(db['Vouchers'])
        return render_template('admin.html', name=name, membercount=membercount, admincount=admincount, suppliercount=suppliercount, inventorycount=inventorycount, ordercount=ordercount, foremcount=foremcount, newscount=newscount, vouchercount=vouchercount)
    else:
        flash("UNAUTHORISED ACCESS. LOG IN TO ACCESS", category='error')
        return redirect(url_for('login'))


@app.route("/addadmin", methods=['GET', 'POST'])
def addadmin():
    create_admin_form = CreateAdminForm(request.form)
    if request.method == 'POST' and create_admin_form.validate():
        admin_dict = {}
        db = shelve.open('database.db', 'c')
        try:
            admin_dict = db['Admin']
        except:
            print("Error in retrieving admins from database.db")
        email_list = []
        admin_email_list = []
        member_list = db['Members']
        admin_list = db['Admin']
        existing_ids = set(admin_list.keys())
        missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
        if missing_ids:
            my_key = min(missing_ids)
        else:
            my_key = len(admin_dict) + 1

        for i in member_list:
            email_list.append(member_list[i].get_email())
        for i in admin_list:
            admin_email_list.append(admin_list[i].get_email())
        if create_admin_form.email.data in email_list:
            flash('email already exists. Please choose a different one.', 'error')
        elif create_admin_form.email.data in admin_email_list:
            flash('email already exists. Please choose a different one.', 'error')
        else:
            admin = Admin.Admin(create_admin_form.first_name.data, create_admin_form.last_name.data,
                                create_admin_form.email.data, create_admin_form.password.data)
            admin.set_member_id(my_key)
            admin_dict[my_key] = admin
            db['Admin'] = admin_dict
            db.close()
            return redirect(url_for('admin'))

    return render_template('registeradmin.html', form=create_admin_form)


@app.route('/viewadmins')
def viewadmins():
    admin_dict = {}
    db = shelve.open('database.db', 'r')
    admin_dict = db['Admin']
    db.close()

    admin_list = []
    for key in admin_dict:
        admin = admin_dict.get(key)
        admin_list.append(admin)
    return render_template('viewadmins.html', count=len(admin_list), admin_list=admin_list)


@app.route('/updateadmin/<int:id>/', methods=['GET', 'POST'])
def update_admin(id):
    update_admin_form = CreateAdminForm(request.form)
    if request.method == 'POST' and update_admin_form.validate():
        admin_dict = {}
        db = shelve.open('database.db', 'w')
        admin_dict = db['Admin']
        admin = admin_dict.get(id)

        admin.set_first_name(update_admin_form.first_name.data)
        admin.set_last_name(update_admin_form.last_name.data)
        admin.set_email(update_admin_form.email.data)
        admin.set_password(update_admin_form.password.data)

        db['Admin'] = admin_dict
        db.close()
        return redirect(url_for('viewadmins'))
    else:
        admin_dict = {}
        db = shelve.open('database.db', 'w')
        admin_dict = db['Admin']
        db.close()

        admin = admin_dict.get(id)
        update_admin_form.first_name.data = admin.get_first_name()
        update_admin_form.last_name.data = admin.get_last_name()
        update_admin_form.email.data = admin.get_email()
        update_admin_form.password.data = admin.get_password()
        return render_template('updateAdmin.html', form=update_admin_form)


@app.route('/deleteadmin/<int:id>/', methods=['POST'])
def delete_admin(id):
    admin_dict = {}
    db = shelve.open('database.db', 'w')
    admin_dict = db['Admin']
    admin_dict.pop(id)
    db['Admin'] = admin_dict
    db.close()
    return redirect(url_for('viewadmins'))


@app.route('/register', methods=['GET', 'POST'])
def create_member():
    create_member_form = CreateMemberForm(request.form)
    if request.method == 'POST' and create_member_form.validate():
        members_dict = {}
        db = shelve.open('database.db', 'c')
        try:
            members_dict = db['Members']
        except:
            print("Error in retrieving members from database.db")

        email_list = []
        admin_email_list = []
        member_list = db['Members']
        admin_list = db['Admin']

        existing_ids = set(member_list.keys())
        missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
        if missing_ids:
            my_key = min(missing_ids)
        else:
            my_key = len(members_dict) + 1

        for i in member_list:
            email_list.append(member_list[i].get_email())
        for i in admin_list:
            admin_email_list.append(admin_list[i].get_email())
        if create_member_form.email.data in email_list:
            flash('email already exists. Please choose a different one.', 'error')
        elif create_member_form.email.data in admin_email_list:
            flash('email already exists. Please choose a different one.', 'error')
        else:
            member = Member.Member(create_member_form.first_name.data, create_member_form.last_name.data,
                                   create_member_form.birthdate.data, create_member_form.gender.data,
                                   create_member_form.email.data, create_member_form.phone.data,
                                   create_member_form.password.data)
            member.set_member_id(my_key)
            members_dict[my_key] = member
            db['Members'] = members_dict
            db.close()
            return redirect(url_for('registrationconfirmation'))
    return render_template('adminmembers.html', form=create_member_form)


@app.route('/registrationconfirmation')
def registrationconfirmation():
    return render_template('registrationconfirmation.html')


@app.route('/members')
def members():
    return render_template('adminmembers.html')


@app.route('/members/viewmembers', methods=['GET', 'POST'])
def view_members():
    create_search_form = CreateSearchForm(request.form)
    member_list = []

    if request.method == 'POST' and create_search_form.validate():
        search = create_search_form.search.data.lower()

        with shelve.open('database.db', 'r') as db:
            member_dict = db['Members']

            try:
                search_int = int(search)

                matching_age_members = [member for member_id, member in member_dict.items() if
                                        member.get_age() == search_int]
                matching_phone_members = [member for member_id, member in member_dict.items() if
                                          member.get_phone() == search]

                member_list = matching_age_members + matching_phone_members
            except ValueError:
                matching_first_name_members = [member for member_id, member in member_dict.items() if
                                               member.get_first_name().lower() == search]
                matching_last_name_members = [member for member_id, member in member_dict.items() if
                                              member.get_last_name().lower() == search]
                matching_gender_members = [member for member_id, member in member_dict.items() if
                                           member.get_gender().lower() == search]
                matching_email_members = [member for member_id, member in member_dict.items() if
                                          member.get_email().lower() == search]
                matching_phone_members = [member for member_id, member in member_dict.items() if
                                          member.get_phone() == search]
                matching_passwords = [member for member_id, member in member_dict.items() if
                                      member.get_password().lower() == search]

                if matching_first_name_members:
                    member_list = matching_first_name_members
                elif matching_last_name_members:
                    member_list = matching_last_name_members
                elif matching_gender_members:
                    member_list = matching_gender_members
                elif matching_email_members:
                    member_list = matching_email_members
                elif matching_phone_members:
                    member_list = matching_phone_members
                elif matching_passwords:
                    member_list = matching_passwords

    else:
        with shelve.open('database.db', 'r') as db:
            member_dict = db['Members']
            member_list = list(member_dict.values())

    return render_template('adminviewmembers.html', count=len(member_list), member_list=member_list,
                           form=create_search_form)


@app.route('/updatemember/<int:id>/', methods=['GET', 'POST'])
def update_user(id):
    update_member_form = CreateMemberForm(request.form)
    if request.method == 'POST' and update_member_form.validate():
        members_dict = {}
        db = shelve.open('database.db', 'w')
        members_dict = db['Members']
        member = members_dict.get(id)
        member.set_first_name(update_member_form.first_name.data)
        member.set_last_name(update_member_form.last_name.data)
        member.set_email(update_member_form.email.data)
        member.set_phone(update_member_form.phone.data)
        member.set_password(update_member_form.password.data)
        db['Members'] = members_dict
        db.close()
        return redirect(url_for('view_members'))
    else:
        members_dict = {}
        db = shelve.open('database.db', 'r')
        members_dict = db['Members']
        db.close()
        member = members_dict.get(id)
        update_member_form.first_name.data = member.get_first_name()
        update_member_form.last_name.data = member.get_last_name()
        update_member_form.email.data = member.get_email()
        update_member_form.phone.data = member.get_phone()
        update_member_form.password.data = member.get_password()
        return render_template('adminupdatemember.html', form=update_member_form)


@app.route('/deletemember/<int:id>', methods=['POST'])
def delete_user(id):
    member_dict = {}
    db = shelve.open('database.db', 'w')
    member_dict = db['Members']
    member_dict.pop(id)
    db['Members'] = member_dict
    db.close()
    return redirect(url_for('view_members'))


@app.route('/addproduct', methods=['GET', 'POST'])
def create_product():
    create_product_form = CreateProductForm(request.form)
    if request.method == 'POST' and create_product_form.validate():
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

            inventory_dict = {}
            db = shelve.open('database.db', 'c')
            try:
                inventory_dict = db['Inventory']
            except:
                print("Error in retrieving products from database.db")

            existing_ids = set(inventory_dict.keys())
            missing_ids = set(range(1, max(existing_ids) + 2)) - existing_ids
            if missing_ids:
                my_key = min(missing_ids)
            else:
                my_key = len(inventory_dict) + 1

            product = Product.Product(create_product_form.name.data, create_product_form.price.data,
                                      create_product_form.category.data, create_product_form.remarks.data,
                                      create_product_form.drinks.data, filename)
            product.set_product_id(my_key)
            inventory_dict[my_key] = product
            db['Inventory'] = inventory_dict
            db.close()

            return redirect(url_for('admin'))
    return render_template('admininventory.html', form=create_product_form)


@app.route('/inventory')
def inventory():
    return render_template('admininventory.html')


@app.route('/inventory/viewinventory', methods=['GET', 'POST'])
def view_inventory():
    create_search_form = CreateSearchForm(request.form)
    inventory_list = []

    if request.method == 'POST' and create_search_form.validate():
        search = create_search_form.search.data.lower()

        with shelve.open('database.db', 'r') as db:
            inventory_dict = db['Inventory']

            try:
                search_int = int(search)

                matching_price_product = [product for product_id, product in inventory_dict.items() if
                                            str(product.get_price()) == search]  # Compare as string

                inventory_list = matching_price_product
            except ValueError:
                matching_name_product = [product for product_id, product in inventory_dict.items() if
                                           product.get_name().lower() == search]
                matching_category_product = [product for product_id, product in inventory_dict.items() if
                                         product.get_category().lower() == search]
                matching_remarks_product = [product for product_id, product in inventory_dict.items() if
                                         product.get_remarks().lower() == search]
                matching_drinks_product = [product for product_id, product in inventory_dict.items() if
                                         product.get_drinks().lower() == search]

                inventory_list = matching_name_product + matching_category_product + matching_remarks_product + matching_drinks_product

    else:
        with shelve.open('database.db', 'r') as db:
            inventory_dict = db['Inventory']
            inventory_list = list(inventory_dict.values())

    return render_template('adminviewinventory.html', count=len(inventory_list), inventory_list=inventory_list,
                           form=create_search_form)

@app.route('/updateproduct/<int:id>/', methods=['GET', 'POST'])
def update_product(id):
    update_product_form = CreateProductForm(request.form)
    if request.method == 'POST' and update_product_form.validate():
        inventory_dict = {}
        db = shelve.open('database.db', 'w')
        inventory_dict = db['Inventory']
        product = inventory_dict.get(id)
        product.set_name(update_product_form.name.data)
        product.set_price(update_product_form.price.data)
        product.set_category(update_product_form.category.data)
        product.set_remarks(update_product_form.remarks.data)
        product.set_drinks(update_product_form.drinks.data)

        if 'file' in request.files:
            file = request.files['file']

            if file.filename != '':
                if allowed_file(file.filename):
                    old_filename = product.get_image()
                    old_filepath = os.path.join(app.config['UPLOAD_FOLDER'], old_filename)
                    if os.path.exists(old_filepath):
                        os.remove(old_filepath)

                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    product.set_image(filename)

        db['Inventory'] = inventory_dict
        db.close()
        return redirect(url_for('view_inventory'))
    else:
        inventory_dict = {}
        db = shelve.open('database.db', 'r')
        inventory_dict = db['Inventory']
        db.close()
        product = inventory_dict.get(id)
        update_product_form.name.data = product.get_name()
        update_product_form.price.data = product.get_price()
        update_product_form.category.data = product.get_category()
        update_product_form.remarks.data = product.get_remarks()
        update_product_form.drinks.data = product.get_drinks()
        return render_template('adminupdateproduct.html', form=update_product_form)


@app.route('/deleteproduct/<int:id>', methods=['POST'])
def delete_product(id):
    inventory_dict = {}
    db = shelve.open('database.db', 'w')
    inventory_dict = db['Inventory']
    inventory_dict.pop(id)
    db['Inventory'] = inventory_dict
    db.close()
    return redirect(url_for('view_inventory'))


# Forum vvv


@app.route('/createQuestion', methods=['GET', 'POST'])
def create_question():
    date = datetime.date.today().strftime('%d-%m-%Y')
    create_question_form = CreateQuestionForm(request.form)
    # need to get data from  create_user_form and save to shelve database
    if request.method == 'POST' and create_question_form.validate():
        question = Question(create_question_form.email.data,
                            create_question_form.title.data,
                            create_question_form.question.data,
                            date,
                            create_question_form.overall.data,
                            create_question_form.feedback.data, reply="")
        add_question(question)
        return redirect(url_for('feedbackformconfirmation'))
    return render_template('createQuestion.html', form=create_question_form)


@app.route('/feedbackformconfirmation')
def feedbackformconfirmation():
    return render_template('feedbackformconfirmation.html')


@app.route('/viewQuestion')
def retrieve_questions():
    questions_dict = {}
    db = shelve.open('database.db', 'r')
    questions_dict = db['Question']
    db.close()
    questions_list = []
    for key in questions_dict:
        question = questions_dict.get(key)
        questions_list.append(question)
    return render_template('retrieveQuestion.html', count=len(questions_list), questions_list=questions_list)

@app.route('/filter_questions', methods=['GET', 'POST'])
def filter_questions():
    value = request.args.get('filter_type') or request.form.get('filter_type')

    db = shelve.open('database.db', 'r')
    questions_dict = db['Question']
    db.close()

    if value == 'all':
        questions_list = list(questions_dict.values())

    elif value == 'question':
        questions_list = [question for question in questions_dict.values() if
                          question.get_question() != "" and question.get_feedback() == ""]

    elif value == "feedback":
        questions_list = [question for question in questions_dict.values() if
                          question.get_feedback() != "" and question.get_question() == ""]

    elif value == "both":
        questions_list = [question for question in questions_dict.values() if
                          question.get_feedback() != "" and question.get_question() != ""]

    elif value == "reply":
        questions_list = [question for question in questions_dict.values() if question.get_reply() != ""]

    elif value == "no_reply":
        questions_list = [question for question in questions_dict.values() if question.get_reply() == ""]

    return render_template('retrieveQuestion.html', count=len(questions_list), questions_list=questions_list)

    
@app.route('/cviewQuestion')
def cretrieve_questions():
    questions_dict = {}
    db = shelve.open('database.db', 'r')
    questions_dict = db['Question']
    db.close()
    questions_list = []
    for key in questions_dict:
        question = questions_dict.get(key)
        questions_list.append(question)
    return render_template('customerforum.html', count=len(questions_list), questions_list=questions_list)


@app.route('/updateQuestion/<int:id>/', methods=['GET', 'POST'])
def update_question(id):
    update_question_form = CreateReplyForm(request.form)

    if request.method == 'POST' and update_question_form.validate():
        print("updating question")
        db = shelve.open('database.db', 'w')
        questions_dict = db['Question']
        question = questions_dict.get(id)
        question.set_reply(update_question_form.reply.data)
        db['Question'] = questions_dict
        db.close()
        return redirect(url_for('retrieve_questions'))

    else:
        db = shelve.open('database.db', 'r')
        questions_dict = db['Question']
        db.close()
        question = questions_dict.get(id)
        id = question.get_question_id()
        email = question.get_email()
        title = question.get_title()
        question_text = question.get_question()
        overall = question.get_overall()
        feedback = question.get_feedback()

        rdate = datetime.date.today()
        reply = question.get_reply()

        update_question_form.reply.data = reply
        date = question.get_date_posted()
        return render_template('updateQuestion.html', form=update_question_form, date=date, rdate=rdate, id=id,email=email, title=title,
                               question=question_text, overall=overall, feedback=feedback)


@app.route('/deleteQuestion/<int:id>', methods=['POST'])
def delete_questions(id):
    questions_dict = {}
    db = shelve.open('database.db', 'w')
    questions_dict = db['Question']

    questions_dict.pop(id)

    db['Question'] = questions_dict
    db.close()
    return redirect(url_for('retrieve_questions'))


@app.route('/filter_outbox', methods=['GET', 'POST'])
def filter_outbox():
    # Retrieve products from inventory
    inventory_dict = {}
    db_inventory = shelve.open('database.db', 'r')
    inventory_dict = db_inventory['Inventory']
    db_inventory.close()

    categories = set(product.get_category() for product in inventory_dict.values())

    selected_category = request.form.get('category')  # Change to request.form for POST
    if selected_category:
        filtered_products = {k: v for k, v in inventory_dict.items() if v.get_category() == selected_category}
    else:
        filtered_products = inventory_dict

    return render_template('outbox.html', outbox_products=filtered_products.values(), categories=categories)


@app.route('/createvoucher', methods=['GET', 'POST'])
def createvoucher():
    create_voucher_form = CreateVoucherForm(request.form)
    if request.method == 'POST' and create_voucher_form.validate():
        voucher_dict = {}
        db = shelve.open('database.db', 'c')
        try:
            voucher_dict = db['Vouchers']
        except:
            print("Error in retrieving vouchers from database.db")

        voucher_list = []
        for i in voucher_dict:
            voucher_list.append(voucher_dict[i].get_voucher_id())
        if create_voucher_form.voucher_id.data in voucher_list:
            flash('voucher already exists.', 'error')
        else:
            voucher = Voucher.Voucher(create_voucher_form.voucher_id.data,
                                      create_voucher_form.name.data,
                                      create_voucher_form.discount.data)
            voucher_dict[create_voucher_form.voucher_id.data] = voucher
            db['Vouchers'] = voucher_dict
            db.close()
            return redirect(url_for('admin'))
    return render_template('createVoucher.html', form=create_voucher_form)


@app.route('/viewvouchers')
def viewvouchers():
    voucher_dict = {}
    db = shelve.open('database.db', 'r')
    voucher_dict = db['Vouchers']
    db.close()

    voucher_list = []

    for key in voucher_dict:
        voucher = voucher_dict.get(key)
        voucher_list.append(voucher)

    return render_template('viewVouchers.html', count=len(voucher_list), voucher_list=voucher_list)


@app.route('/deletevoucher/<id>', methods=['POST'])
def delete_voucher(id):
    voucher_dict = {}
    db = shelve.open('database.db', 'w')
    voucher_dict = db['Vouchers']
    member_dict = db['Members']
    for i in member_dict:
        vouchers = member_dict[i].get_vouchers()
        if voucher_dict[id].get_voucher_id() in vouchers:
            vouchers.remove(voucher_dict[id].get_voucher_id())
    voucher_dict.pop(id)
    db['Vouchers'] = voucher_dict
    db['Members'] = member_dict
    db.close()
    return redirect(url_for('viewvouchers'))


@app.route('/givevoucher', methods=['GET', 'POST'])
def givevoucher():
    voucher_form = VoucherForm(request.form)
    db = shelve.open('database.db', 'r')
    member_dict = db['Members']
    voucher_dict = db['Vouchers']
    member_list = []
    voucher_list = []
    for i in member_dict:
        member_list.append(member_dict[i])
    for i in voucher_dict:
        voucher_list.append(voucher_dict[i])

    if request.method == "POST" and voucher_form.validate():
        db = shelve.open('database.db', 'w')
        member_dict = db['Members']
        voucher_dict = db['Vouchers']
        email_list = []
        voucher_list = []

        for i in db['Members']:
            email_list.append(member_dict[i].get_email())
        for i in voucher_dict:
            voucher_list.append(voucher_dict[i].get_voucher_id())

        if voucher_form.email.data not in email_list:
            flash('Email does not exist', 'error')
        elif voucher_form.voucher_id.data not in voucher_list:
            flash('Voucher does not exist', 'error')
        else:
            for i in member_dict:
                if member_dict[i].get_email() == voucher_form.email.data:
                    member_dict[i].set_vouchers(voucher_form.voucher_id.data)
                    flash(f'coupon successfully given to {voucher_form.email.data}')
                    break
            db['Members'] = member_dict
        voucher_list = []
        for i in voucher_dict:
            voucher_list.append(voucher_dict[i])
        db.close()
    return render_template("givevoucher.html", form=voucher_form, member_list=member_list, voucher_list=voucher_list)


@app.route('/download_excel/<db_name>')
def excel_converter(db_name):
    try:
        with shelve.open('database.db', 'r') as db:
            data_dict = db.get(db_name, {})
    except Exception as e:
        print(f"Error in retrieving data from database.db: {e}")
        return

    if not data_dict:
        print(f"No data found for '{db_name}' in the database.")
        return

    # convert data to DataFrame
    df = pd.DataFrame([obj.as_dict() for obj in data_dict.values()])

    # new Excel workbook
    wb = Workbook()
    ws = wb.active

    # write DataFrame to Excel
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # set 'wrap text' for all cells
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True)

    # 'Image' refers to the image file name written in text
    # 'Image file' refers to the actual picture in png, jpg, jpeg, or gif

    # add images (if 'Image' column exists)
    if 'Image' in df.columns:
        image_folder = 'static/uploads'
        for index, row in df.iterrows():
            image_filename = row['Image']
            if image_filename:
                # os.path.join method to construct the file path, and os.path.normpath method to normalize the path
                image_path = os.path.normpath(os.path.join(image_folder, image_filename))
                img = XLImage(image_path)
                # img.width = 120  # no need to adjust the image dimensions for now
                # img.height = 160
                ws.add_image(img, f'H{index + 2}')  # 'Image file' column to be in column H
                ws['H1'] = 'Image file'  # add title

    # save
    excel_filename = f'{db_name}.xlsx'
    try:
        wb.save(excel_filename)
        print(f"Excel file '{excel_filename}' created successfully.")
        return send_file(excel_filename, as_attachment=True)
    except Exception as e:
        print(f"Error saving Excel file '{excel_filename}': {e}")


@app.route('/beanbox')
def beanbox():
    return render_template('beanbox.html')


@app.route('/feedbackreport')
def feedback_report():
    df = combine_databases()

    # # convert 'Question_Date Posted' to DateTime format
    # df['Question_Date Posted'] = pd.to_datetime(df['Question_Date Posted'])
    #
    # # add month columns into df
    # df['Question_Date Posted_Month'] = df['Question_Date Posted'].dt.month

    positive_reviews = (df['Question_Overall rating'] == "E").sum()
    total_reviews = df['Question_Overall rating'].isin(["E", "N", "B"]).sum()
    positive_percentage = positive_reviews / total_reviews

    # liquid chart for % of positive reviews
    overall_positive = (
        Liquid()
        .add("Positive Reviews", [positive_percentage, positive_reviews, 0.15, 0.05])
        # .add("Positive Reviews", [0.68, 0.60, 0.15, 0.05])
        .set_global_opts(title_opts=opts.TitleOpts(title="% of Positive Reviews"))
    )

    # create a Tab and add the chart for the percentage of positive reviews
    tab = Tab(page_title='Feedback Overview')
    tab.add(overall_positive, 'Percentage of Positive Reviews')

    # render the tab directly in the HTML template
    chart_html = Markup(tab.render_embed())

    return render_template('feedbackreport.html', chart_html=chart_html)


@app.route('/performancereport')
def performance_report():
    df = combine_databases()

    # convert 'OrderHist_Date' to DateTime format
    df['OrderHist_Date'] = pd.to_datetime(df['OrderHist_Date'])

    # add month columns into df
    df['OrderHist_Month'] = df['OrderHist_Date'].dt.month

    # group df by months and sum the 'OrderHist_Payment amount'
    grouped_by_months = df.groupby(by=['OrderHist_Month'])['OrderHist_Payment amount'].sum()

    # create Estimated Profit (80% of Sales)
    estimated_profit = grouped_by_months * 0.8

    # plot Sales (Payment amount) and Estimated Profit by months
    bar_chart_by_month = (
        Bar()
        .add_xaxis(grouped_by_months.index.tolist())
        .add_yaxis('Sales', grouped_by_months.round().tolist())
        .add_yaxis('Estimated Profit', estimated_profit.round().tolist())
        .set_global_opts(
            title_opts=opts.TitleOpts(title='Sales and Profit by Month', subtitle='in SGD($)')
        )
    )

    bar_chart_by_month.render_notebook()

    # create a Tab and add the chart for monthly overview
    tab = Tab(page_title='Sales Overview')
    tab.add(bar_chart_by_month, 'Sales and Profit by Month')

    # group data by product category and sum 'OrderHist_Payment amount'
    grouped_by_product_category = df.groupby(by='Inventory_Category', as_index=False)['OrderHist_Payment amount'].sum()

    # create Estimated Profit (80% of Sales) for each product category
    grouped_by_product_category['Estimated Profit'] = grouped_by_product_category['OrderHist_Payment amount'] * 0.8

    # plot Sales (Payment amount) and Estimated Profit by product category
    bar_chart_by_category = (
        Bar()
        .add_xaxis(grouped_by_product_category['Inventory_Category'].tolist())
        .add_yaxis('Sales', grouped_by_product_category['OrderHist_Payment amount'].round().tolist())
        .add_yaxis('Estimated Profit', grouped_by_product_category['Estimated Profit'].round().tolist())
        .reversal_axis()
        .set_series_opts(label_opts=opts.LabelOpts(position='right'))
        .set_global_opts(
            title_opts=opts.TitleOpts(title='Product Category', subtitle='in SGD($)')
        )
    )

    bar_chart_by_category.render_notebook()

    # create a Tab and add the chart for product category overview
    tab.add(bar_chart_by_category, 'Sales and Profit by Product Category')

    # render the tab directly in the HTML template
    chart_html = Markup(tab.render_embed())

    return render_template('performancereport.html', chart_html=chart_html)

@app.route('/view_more/<int:product_id>')
def view_more(product_id):
    inventory_dict = {}
    db_inventory = shelve.open('database.db', 'r')
    inventory_dict = db_inventory.get('Inventory', {})
    db_inventory.close()

    selected_product = inventory_dict.get(product_id)

    return render_template('view_more.html', selected_product=selected_product)

@app.route('/supplier', methods=['GET', 'POST'])
def create_supplier():
    create_supplier_form = CreateSupplierForm(request.form)
    if request.method == 'POST' and create_supplier_form.validate():
        supplier_dict = {}
        with shelve.open('database.db', 'c') as db:
            try:
                supplier_dict = db['Supplier']
            except:
                print("Error in retrieving members from database.db")

            email_list = []
            for i in supplier_dict:
                email_list.append(supplier_dict[i].get_company_email())
            if create_supplier_form.company_email.data in email_list:
                flash('email already exists. Please choose a different one.', 'error')
            else:
                supplier = Supplier.Supplier(create_supplier_form.company_name.data, create_supplier_form.company_email.data,
                                             create_supplier_form.company_phone.data, create_supplier_form.company_address.data,
                                             create_supplier_form.password.data)
                if len(supplier_dict) == 0:
                    my_key = 1
                else:
                    my_key = len(supplier_dict.keys()) + 1
                supplier.set_supplier_id(my_key)
                supplier_dict[my_key] = supplier
                db['Supplier'] = supplier_dict
                db.close()
            return redirect(url_for('registrationconfirmation'))
    return render_template('adminsuppliers.html', form=create_supplier_form)
@app.route('/supplier/viewsuppliers', methods=['GET', 'POST'])
def view_suppliers():
    create_search_form = CreateSearchForm(request.form)
    supplier_list = []

    if request.method == 'POST' and create_search_form.validate():
        search = create_search_form.search.data.lower()

        with shelve.open('database.db', 'r') as db:
            supplier_dict = db['Supplier']

            try:
                search_int = int(search)  # Try converting search input to integer

                matching_phone_suppliers = [supplier for supplier_id, supplier in supplier_dict.items() if
                                            str(supplier.get_company_phone()) == search]  # Compare as string

                supplier_list = matching_phone_suppliers
            except ValueError:
                # Handle non-integer search inputs as string searches
                matching_name_suppliers = [supplier for supplier_id, supplier in supplier_dict.items() if
                                           supplier.get_company_name().lower() == search]
                matching_email_suppliers = [supplier for supplier_id, supplier in supplier_dict.items() if
                                            supplier.get_company_email().lower() == search]
                matching_address_suppliers = [supplier for supplier_id, supplier in supplier_dict.items() if
                                              supplier.get_company_address().lower() == search]
                matching_passwords = [supplier for supplier_id, supplier in supplier_dict.items() if
                                      supplier.get_password().lower() == search]

                # Combine all matches
                supplier_list = matching_name_suppliers + matching_email_suppliers + matching_address_suppliers + matching_passwords

    else:
        with shelve.open('database.db', 'r') as db:
            supplier_dict = db['Supplier']
            supplier_list = list(supplier_dict.values())

    return render_template('adminviewsuppliers.html', count=len(supplier_list), supplier_list=supplier_list,
                           form=create_search_form)

@app.route('/updatesupplier/<int:id>/', methods=['GET', 'POST'])
def update_supplier(id):
    update_supplier_form = CreateSupplierForm(request.form)
    if request.method == 'POST' and update_supplier_form.validate():
        supplier_dict = {}
        db = shelve.open('database.db', 'w')
        supplier_dict = db['Supplier']
        supplier = supplier_dict.get(id)
        supplier.set_company_name(update_supplier_form.company_name.data)
        supplier.set_company_email(update_supplier_form.company_email.data)
        supplier.set_company_phone(update_supplier_form.company_phone.data)
        supplier.set_company_address(update_supplier_form.company_address.data)
        supplier.set_password(update_supplier_form.password.data)
        db['Supplier'] = supplier_dict
        db.close()
        return redirect(url_for('view_suppliers'))
    else:
        supplier_dict = {}
        db = shelve.open('database.db', 'r')
        supplier_dict = db['Supplier']
        db.close()
        supplier = supplier_dict.get(id)
        update_supplier_form.company_name.data = supplier.get_company_name()
        update_supplier_form.company_email.data = supplier.get_company_email()
        update_supplier_form.company_phone.data = supplier.get_company_phone()
        update_supplier_form.company_address.data = supplier.get_company_address()
        update_supplier_form.password.data = supplier.get_password()
        return render_template('adminupdatesupplier.html', form=update_supplier_form)

@app.route('/deletesupplier/<int:id>', methods=['POST'])
def delete_supplier(id):
    supplier_dict = {}
    db = shelve.open('database.db', 'w')
    supplier_dict = db['Supplier']
    supplier_dict.pop(id)
    db['Supplier'] = supplier_dict
    db.close()
    return redirect(url_for('view_suppliers'))


@app.route('/createNews', methods=['GET', 'POST'])
def create_news():
    date = datetime.date.today().strftime('%d-%m-%Y')
    create_news_form = CreateNewsForm(request.form)

    if request.method == 'POST' and create_news_form.validate():
        if 'file' not in request.files:
            flash("No file part")
            return redirect(request.url)
        file = request.files["file"]
        if file.filename == "":
            flash("No image selected for uploading")
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            flash('Image successfully uploaded')
        else:
            flash('Allowed image types are png, jpg, jpeg, gif')
            return redirect(request.url)

        news = News(create_news_form.title.data,
                    create_news_form.description.data,
                    date, filename)

        add_news(news)
        return redirect(url_for('homepage'))
    return render_template('createNews.html', form=create_news_form, date=date)


@app.route('/viewNews')
def retrieve_news():
    news_dict = {}
    db = shelve.open('database.db', 'r')
    news_dict = db['News']
    db.close()
    news_list = []

    for key in news_dict:
        news = news_dict.get(key)
        news_list.append(news)

    return render_template('retrieveNews.html', count=len(news_list), news_list=news_list)


@app.route('/cviewNews')
def cretrieve_news():
    news_dict = {}
    db = shelve.open('database.db', 'r')
    news_dict = db['News']
    db.close()
    news_list = []


    for key in news_dict:
        news = news_dict.get(key)
        news_list.append(news)
    return render_template('customernews.html', count=len(news_list), news_list=news_list)


@app.route('/updateNews/<int:id>/', methods=['GET', 'POST'])
def update_news(id):
    update_news_form = CreateNewsForm(request.form)

    if request.method == 'POST' and update_news_form.validate():
        print("updating news")
        db = shelve.open('database.db', 'w')
        news_dict = db['News']

        news = news_dict.get(id)
        news.set_title(update_news_form.title.data)
        news.set_description(update_news_form.description.data)

        if 'file' in request.files:
            file = request.files['file']

            if file.filename != '':
                if allowed_file(file.filename):
                    old_filename = news.get_file()
                    old_filepath = os.path.join(app.config['UPLOAD_FOLDER'], old_filename)
                    if os.path.exists(old_filepath):
                        os.remove(old_filepath)

                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    news.set_file(filename)

        db['News'] = news_dict
        db.close()
        return redirect(url_for('retrieve_news'))

    else:
        db = shelve.open('database.db', 'r')
        news_dict = db['News']
        db.close()
        news = news_dict.get(id)
        update_news_form.title.data = news.get_title()
        update_news_form.description.data = news.get_description()
        return render_template('updateNews.html', form=update_news_form)


@app.route('/deleteNews/<int:id>', methods=['POST'])
def delete_news(id):
    db = shelve.open('database.db', 'w')
    news_dict = db['News'].copy()

    if id in news_dict:
        news_dict.pop(id)

    db['News'] = news_dict
    db.close()
    return redirect(url_for('retrieve_news'))


if __name__ == '__main__':
    app.run(debug=True)
